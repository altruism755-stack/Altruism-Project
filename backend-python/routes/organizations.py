import csv
import io
import json
import os
import re
import secrets
from datetime import datetime, timedelta, timezone

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import JSONResponse, Response
from pydantic import BaseModel
from typing import Any, Optional

from database import get_db, dict_row, dict_rows
from auth import get_current_user, require_roles, require_approved_org_admin

# Base URL for invite links — set APP_BASE_URL in env for production.
_APP_BASE_URL = os.getenv("APP_BASE_URL", "http://localhost:5173")

_EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")

DUPLICATE_STRATEGIES = {"skip_duplicates", "update_existing", "invite_anyway"}

# Canonical join_source values for analytics/Power BI compatibility.
# Any unrecognized value from CSV import is coerced to "other".
VALID_JOIN_SOURCES = {"website", "referral", "campaign", "other"}

# Incremented when the export schema gains new columns.
# Power BI dashboards can filter on this to detect stale cached exports.
EXPORT_VERSION = "1.1"

# Note: database uses "Active" while the analytics layer uses "accepted".
# Both values are treated as active for consistency across the system.
ACTIVE_STATUSES = {"active", "accepted"}


def _utcnow() -> str:
    """Single source of truth for UTC timestamps across the analytics pipeline.
    Always produces YYYY-MM-DDTHH:MM:SSZ — consistent with ISO 8601 and Power BI."""
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _is_active(status: str) -> int:
    """Derived analytics flag. 1 = active membership, 0 = everything else.
    Case-insensitive and safe against None or empty input."""
    return 1 if (status or "").lower() in ACTIVE_STATUSES else 0


class AddOrgAdminBody(BaseModel):
    email: str


class ImportCSVBody(BaseModel):
    csv: str
    duplicate_strategy: str = "skip_duplicates"


def _make_invite_link(token: str) -> str:
    """Phase 1: returns a link for manual distribution.
    Phase 2: call your email service here and keep this as the single integration point."""
    return f"{_APP_BASE_URL}/accept-invite?token={token}"


def _validate_csv_row(row: dict, row_num: int) -> tuple:
    """Returns (cleaned_row, error_dict). Exactly one will be None."""
    name  = (row.get("name")  or row.get("Name")  or "").strip()
    email = (row.get("email") or row.get("Email") or "").strip().lower()
    if not name:
        return None, {"row": row_num, "reason": "Missing name"}
    if not email:
        return None, {"row": row_num, "reason": "Missing email"}
    if not _EMAIL_RE.match(email):
        return None, {"row": row_num, "reason": f"Invalid email format: {email}"}
    # Prefer explicit join_source; fall back to legacy source column; default to "other".
    raw_js = (row.get("join_source") or row.get("source") or "other").strip().lower()
    join_source = raw_js if raw_js in VALID_JOIN_SOURCES else "other"
    # channel_detail is only meaningful for campaign joins; blank it out otherwise.
    raw_cd = (row.get("channel_detail") or "").strip().lower()
    channel_detail = raw_cd if join_source == "campaign" else ""
    # Location snapshot fields — stored separately so profile edits don't mutate history.
    governorate_snap = (row.get("governorate") or "").strip()
    city_snap        = (row.get("city")        or "").strip()
    return {
        "name":               name,
        "email":              email,
        "phone":              (row.get("phone")      or "").strip(),
        "city":               city_snap or governorate_snap,   # backward-compat combined field
        "skills":             [s.strip() for s in (row.get("skills") or "").split(",") if s.strip()],
        "department":         (row.get("department") or "").strip(),
        "role":               (row.get("role")       or "volunteer").strip(),
        "source":             (row.get("source")     or "manual_import").strip(),
        "join_source":        join_source,
        "channel_detail":     channel_detail,
        "governorate_snap":   governorate_snap,
        "city_snap":          city_snap,
        # Quality-logging marker: non-empty when the original value was coerced.
        "_raw_join_source":   raw_js if raw_js != join_source else "",
    }, None


def _parse_csv(csv_text: str) -> tuple:
    """Parse CSV text. Returns (valid_rows, errors, duplicate_emails_in_csv, quality_warnings).
    quality_warnings: list of rows where join_source was invalid and coerced to 'other'.
    """
    reader           = csv.DictReader(io.StringIO(csv_text.strip()))
    valid_rows       = []
    errors           = []
    seen_emails      = set()
    csv_dupes        = set()
    quality_warnings = []

    for i, row in enumerate(reader, start=2):  # row 1 = header
        cleaned, err = _validate_csv_row(row, i)
        if err:
            errors.append(err)
            continue
        email = cleaned["email"]
        if email in seen_emails:
            # Keep first occurrence; skip subsequent duplicates without raising an error.
            csv_dupes.add(email)
            quality_warnings.append({"row": i, "field": "email", "issue": "duplicate_in_file"})
            continue
        seen_emails.add(email)
        # Surface join_source coercions so callers can log data-quality issues.
        if cleaned.get("_raw_join_source"):
            quality_warnings.append({
                "row": i, "field": "join_source",
                "original": cleaned["_raw_join_source"], "coerced_to": "other",
            })
        valid_rows.append(cleaned)

    return valid_rows, errors, csv_dupes, quality_warnings


# Fields the org admin can update directly on their own profile.
# These are applied immediately without any review step.
EDITABLE_FIELDS = {
    "description", "logo_url", "phone", "website",
    "category", "categories",
    "branches",
    "submitter_name", "submitter_role", "additional_notes",
}
# Fields that require platform-admin review — changes are queued, not applied.
SENSITIVE_FIELDS = {
    "name", "official_email",
    "org_type", "founded_year", "org_size",
    "location",   # HQ governorate
    "hq_city",
}

# Columns whose value is a JSON-encoded array — serialized on write.
JSON_ARRAY_FIELDS = {"branches", "categories"}


class OrgProfileUpdate(BaseModel):
    # Instantly applied fields
    description: Optional[str] = None
    logo_url: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    category: Optional[str] = None        # legacy comma-joined string
    categories: Optional[list] = None
    branches: Optional[list] = None
    submitter_name: Optional[str] = None
    submitter_role: Optional[str] = None
    additional_notes: Optional[str] = None
    # Sensitive — accepted but queued for platform-admin review.
    name: Optional[str] = None
    official_email: Optional[str] = None
    org_type: Optional[str] = None
    founded_year: Optional[str] = None
    org_size: Optional[str] = None
    location: Optional[str] = None        # HQ governorate
    hq_city: Optional[str] = None


def _decode_org_json(org: dict) -> dict:
    """Parse JSON-encoded list columns into native Python lists."""
    if not org:
        return org
    for f in JSON_ARRAY_FIELDS:
        raw = org.get(f)
        if isinstance(raw, str) and raw.strip():
            try:
                org[f] = json.loads(raw)
            except Exception:
                org[f] = []
        elif raw is None:
            org[f] = []
    return org


def _encode_for_db(field: str, value: Any) -> Any:
    if field in JSON_ARRAY_FIELDS:
        return json.dumps(value or [])
    return value


router = APIRouter(prefix="/api/organizations", tags=["organizations"])


# ── /me routes MUST come before /{org_id} so FastAPI doesn't try to cast
# the literal string "me" to an integer and fail with a 422. ─────────────────

def _get_my_org_id(db, user_id: int) -> int:
    org = dict_row(db.execute(
        "SELECT id FROM organizations WHERE admin_user_id = ?", (user_id,)
    ).fetchone())
    if not org:
        raise HTTPException(404, "Organization not found for this user")
    return org["id"]


# ── Organization profile (self-service) ──────────────────────────────────────

@router.get("/me/profile")
def get_my_profile(current_user: dict = Depends(require_approved_org_admin)):
    with get_db() as db:
        org = dict_row(db.execute(
            # Join with users to expose the primary admin's email as submitter_email.
            # admin_user_id IS the submitter — derived from the authoritative users table,
            # so it's correct for all existing orgs without any column migration.
            "SELECT o.*, u.email AS submitter_email "
            "FROM organizations o JOIN users u ON u.id = o.admin_user_id "
            "WHERE o.admin_user_id = ?",
            (current_user["id"],),
        ).fetchone())
        if not org:
            raise HTTPException(404, "Organization not found")
        org = _decode_org_json(org)

        # Forward-compat: if a created_by_user_id column is ever added it takes
        # precedence over admin_user_id as the canonical submitter identity.
        # Today the column doesn't exist so org.get() returns None and the
        # submitter_email set by the JOIN above is used unchanged.
        if org.get("created_by_user_id"):
            override = dict_row(db.execute(
                "SELECT email FROM users WHERE id = ?",
                (org["created_by_user_id"],),
            ).fetchone())
            if override:
                org["submitter_email"] = override["email"]

        pending_rows = dict_rows(db.execute(
            "SELECT id, field, new_value, status, created_at "
            "FROM org_profile_change_requests "
            "WHERE org_id = ? AND status = 'pending' ORDER BY created_at DESC",
            (org["id"],),
        ).fetchall())
        # Build a field → new_value dict for easy frontend consumption.
        pending_dict = {row["field"]: row["new_value"] for row in pending_rows}
        return {
            "current": org,
            "organization": org,          # backward-compat alias
            "pending": pending_dict,
            "pending_changes": pending_rows,
            "has_pending": len(pending_rows) > 0,
        }


@router.put("/me/profile")
def update_my_profile(
    body: OrgProfileUpdate,
    current_user: dict = Depends(require_approved_org_admin),
):
    payload = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}

    with get_db() as db:
        org = dict_row(db.execute(
            "SELECT * FROM organizations WHERE admin_user_id = ?", (current_user["id"],)
        ).fetchone())
        if not org:
            raise HTTPException(404, "Organization not found")

        if not payload:
            org_now = _decode_org_json(dict(org))
            pending = dict_rows(db.execute(
                "SELECT id, field, new_value, status, created_at "
                "FROM org_profile_change_requests "
                "WHERE org_id = ? AND status = 'pending' ORDER BY created_at DESC",
                (org["id"],),
            ).fetchall())
            pending_dict = {row["field"]: row["new_value"] for row in pending}
            return {"organization": org_now, "current": org_now, "applied": [], "queued": [],
                    "pending_changes": pending, "pending": pending_dict,
                    "has_pending": len(pending) > 0, "pending_changes_count": len(pending),
                    "message": "No changes to save."}

        applied: list[str] = []
        queued: list[str] = []

        # Normalize JSON-array fields on the existing row before diffing,
        # so a client sending the same list twice doesn't trigger a write.
        org_decoded = _decode_org_json(dict(org))
        edits = {
            k: v for k, v in payload.items()
            if k in EDITABLE_FIELDS and v != org_decoded.get(k)
        }
        if edits:
            sets = ", ".join(f"{k} = ?" for k in edits)
            values = [_encode_for_db(k, v) for k, v in edits.items()]
            db.execute(
                f"UPDATE organizations SET {sets} WHERE id = ?",
                (*values, org["id"]),
            )
            applied = list(edits.keys())

        for field in SENSITIVE_FIELDS:
            new_val = payload.get(field)
            if new_val is None or new_val == org.get(field):
                continue
            db.execute(
                "DELETE FROM org_profile_change_requests "
                "WHERE org_id = ? AND field = ? AND status = 'pending'",
                (org["id"], field),
            )
            db.execute(
                "INSERT INTO org_profile_change_requests (org_id, requested_by, field, new_value) "
                "VALUES (?, ?, ?, ?)",
                (org["id"], current_user["id"], field, new_val),
            )
            queued.append(field)

        org_after = _decode_org_json(dict_row(db.execute(
            "SELECT * FROM organizations WHERE id = ?", (org["id"],)
        ).fetchone()))
        pending = dict_rows(db.execute(
            "SELECT id, field, new_value, status, created_at "
            "FROM org_profile_change_requests "
            "WHERE org_id = ? AND status = 'pending' ORDER BY created_at DESC",
            (org["id"],),
        ).fetchall())

        if queued and applied:
            message = "Changes saved. Updates to sensitive fields require review."
        elif queued:
            message = "This change requires review."
        elif applied:
            message = "Profile updated."
        else:
            message = "No changes to save."

        pending_dict = {row["field"]: row["new_value"] for row in pending}
        return {
            "organization": org_after,
            "current": org_after,
            "applied": applied,
            "queued": queued,
            "pending_changes": pending,
            "pending": pending_dict,
            "has_pending": len(pending) > 0,
            "pending_changes_count": len(pending),
            "message": message,
        }


# ── Organization admin management (org-scoped) ───────────────────────────────

def _fetch_org_admins(db, org_id: int) -> list:
    return dict_rows(db.execute(
        "SELECT oa.id, oa.user_id, u.email, oa.created_at, oa.role "
        "FROM org_admins oa JOIN users u ON oa.user_id = u.id "
        "WHERE oa.org_id = ? ORDER BY oa.created_at ASC",
        (org_id,),
    ).fetchall())


@router.get("/me/admins")
def list_my_org_admins(current_user: dict = Depends(require_approved_org_admin)):
    with get_db() as db:
        org_id = _get_my_org_id(db, current_user["id"])
        rows = _fetch_org_admins(db, org_id)
        # This acts as a one-time lazy migration for legacy organizations.
        # INSERT OR IGNORE prevents duplicate rows against the UNIQUE(user_id, org_id) constraint.
        if not rows:
            db.execute(
                "INSERT OR IGNORE INTO org_admins (user_id, org_id, role) VALUES (?, ?, 'creator')",
                (current_user["id"], org_id),
            )
            rows = _fetch_org_admins(db, org_id)
        return {"admins": rows}


@router.post("/me/admins", status_code=201)
def add_my_org_admin(body: AddOrgAdminBody, current_user: dict = Depends(require_approved_org_admin)):
    with get_db() as db:
        org_id = _get_my_org_id(db, current_user["id"])
        user = dict_row(db.execute("SELECT id, email FROM users WHERE email = ?", (body.email,)).fetchone())
        if not user:
            raise HTTPException(404, f"No user found with email '{body.email}'")
        # Idempotent check FIRST — handles both "already-admin other" and "already-admin self".
        existing = db.execute(
            "SELECT id FROM org_admins WHERE user_id = ? AND org_id = ?", (user["id"], org_id)
        ).fetchone()
        if existing:
            return JSONResponse(status_code=200, content={"message": f"{body.email} is already an organization admin"})
        # Self-grant prevention: a non-admin user must not be able to escalate their own privileges.
        if user["id"] == current_user["id"]:
            raise HTTPException(403, "You cannot grant yourself admin access")
        # INSERT OR IGNORE defends against any concurrent race to the UNIQUE constraint.
        db.execute("INSERT OR IGNORE INTO org_admins (user_id, org_id, role) VALUES (?, ?, 'admin')", (user["id"], org_id))
        return {"message": f"{body.email} is now an organization admin"}


@router.delete("/me/admins/{admin_id}")
def remove_my_org_admin(admin_id: int, current_user: dict = Depends(require_approved_org_admin)):
    with get_db() as db:
        org_id = _get_my_org_id(db, current_user["id"])
        row = dict_row(db.execute(
            "SELECT id, user_id FROM org_admins WHERE id = ? AND org_id = ?", (admin_id, org_id)
        ).fetchone())
        if not row:
            raise HTTPException(404, "Organization admin not found")
        db.execute("DELETE FROM org_admins WHERE id = ?", (admin_id,))
        return {"message": "Organization admin removed"}


# ── Analytics CSV export (Power BI / flat structure) ─────────────────────────

@router.get("/me/volunteers/export")
def export_volunteers_csv(current_user: dict = Depends(require_approved_org_admin)):
    """Flat, denormalized CSV export for analytics tools (Power BI, etc.).

    Schema v1.1 column order (analytics-optimised):
      export_version, volunteer_id, org_id, email, name,
      join_source, channel_detail, status, is_active,
      joined_at, activity_id, governorate, city

    status values: applied | accepted | inactive | unknown
    is_active: 1 when status = 'accepted', 0 for all other states.
    channel_detail is non-empty only when join_source = 'campaign'.
    joined_at: full ISO 8601 UTC (YYYY-MM-DDTHH:MM:SSZ).
    All nulls replaced with empty strings or 0 for Power BI compatibility.
    """
    with get_db() as db:
        org_id = _get_my_org_id(db, current_user["id"])
        rows = dict_rows(db.execute(
            """
            SELECT
                v.id                                                        AS volunteer_id,
                ov.org_id,
                LOWER(TRIM(v.email))                                        AS email,
                v.name,
                COALESCE(ov.join_source, 'other')                          AS join_source,
                CASE
                    WHEN ov.join_source = 'campaign'
                    THEN LOWER(TRIM(COALESCE(ov.channel_detail, '')))
                    ELSE ''
                END                                                         AS channel_detail,
                CASE ov.status
                    WHEN 'Pending'  THEN 'applied'
                    WHEN 'Active'   THEN 'accepted'
                    WHEN 'Inactive' THEN 'inactive'
                    ELSE            'unknown'
                END                                                         AS status,
                COALESCE(ov.is_active, 0)                                  AS is_active,
                COALESCE(
                    ov.joined_at,
                    ov.joined_date || 'T00:00:00Z',
                    ''
                )                                                           AS joined_at,
                ''                                                          AS activity_id,
                COALESCE(ov.governorate_snapshot, '')                       AS governorate,
                COALESCE(ov.city_snapshot,        '')                       AS city
            FROM volunteers v
            JOIN org_volunteers ov ON v.id = ov.volunteer_id
            WHERE ov.org_id = ?
            ORDER BY COALESCE(ov.joined_at, ov.joined_date) ASC
            """,
            (org_id,),
        ).fetchall())

    fieldnames = [
        "export_version", "volunteer_id", "org_id", "email", "name",
        "join_source", "channel_detail", "status", "is_active",
        "joined_at", "activity_id", "governorate", "city",
    ]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    for row in rows:
        row["export_version"] = EXPORT_VERSION
        writer.writerow(row)

    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="volunteers_{org_id}_v{EXPORT_VERSION}.csv"'},
    )


# ── Full-profile export (Excel / CSV) ────────────────────────────────────────

_EXPORT_COLUMNS = [
    "volunteer_id", "name", "email", "phone", "national_id", "nationality",
    "gender", "date_of_birth", "governorate", "city",
    "education_level", "university_name", "faculty", "field_of_study", "study_year",
    "languages", "skills", "cause_areas", "availability",
    "hours_per_week", "prior_experience", "prior_org", "experiences", "about_me",
    "department", "supervisor_name",
    "membership_status", "join_source", "channel_detail", "joined_at", "is_active",
]

# Columns that store JSON-encoded arrays in the DB → joined with "; " for spreadsheets.
_JSON_LIST_COLUMNS = {"languages", "skills", "cause_areas", "availability", "experiences"}

# Columns that should be rendered as text (preserve leading zeros, Arabic, etc.)
_TEXT_COLUMNS = {"phone", "national_id"}

# Long free-text columns — wider width + wrap in Excel.
_WIDE_WRAP_COLUMNS = {"about_me", "experiences", "skills", "cause_areas", "languages", "availability"}


def _decode_list(raw: Any) -> str:
    if raw in (None, "", 0):
        return ""
    if isinstance(raw, str):
        try:
            parsed = json.loads(raw)
        except Exception:
            return raw
    else:
        parsed = raw
    if isinstance(parsed, list):
        parts = []
        for item in parsed:
            if isinstance(item, dict):
                # experiences: prefer a readable summary
                summary = item.get("title") or item.get("role") or item.get("name") or ""
                org = item.get("organization") or item.get("org") or ""
                if summary and org:
                    parts.append(f"{summary} @ {org}")
                elif summary:
                    parts.append(summary)
                else:
                    parts.append(json.dumps(item, ensure_ascii=False))
            else:
                parts.append(str(item))
        return "; ".join(p for p in parts if p)
    return str(parsed)


def _shape_row(row: dict) -> dict:
    out: dict = {}
    for col in _EXPORT_COLUMNS:
        val = row.get(col)
        if col in _JSON_LIST_COLUMNS:
            out[col] = _decode_list(val)
        elif col == "prior_experience":
            out[col] = "Yes" if val else "No"
        elif col == "is_active":
            out[col] = 1 if val else 0
        elif col == "joined_at":
            out[col] = val or row.get("joined_date") or ""
        elif val is None:
            out[col] = ""
        else:
            out[col] = val
    return out


def _fetch_export_rows(db, org_id: int) -> list[dict]:
    return dict_rows(db.execute(
        """
        SELECT
            v.id AS volunteer_id,
            v.name, v.email, v.phone, v.national_id, v.nationality,
            v.gender, v.date_of_birth, v.governorate, v.city,
            v.education_level, v.university_name, v.faculty, v.field_of_study, v.study_year,
            v.languages, v.skills, v.cause_areas, v.availability,
            v.hours_per_week, v.prior_experience, v.prior_org, v.experiences, v.about_me,
            ov.department,
            s.name AS supervisor_name,
            ov.status AS membership_status,
            COALESCE(ov.join_source, 'other') AS join_source,
            COALESCE(ov.channel_detail, '') AS channel_detail,
            COALESCE(ov.joined_at, ov.joined_date) AS joined_at,
            ov.joined_date,
            COALESCE(ov.is_active, 0) AS is_active
        FROM volunteers v
        JOIN org_volunteers ov ON v.id = ov.volunteer_id
        LEFT JOIN supervisors s ON ov.supervisor_id = s.id
        WHERE ov.org_id = ?
        ORDER BY v.name COLLATE NOCASE ASC
        """,
        (org_id,),
    ).fetchall())


def _safe_filename(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", (name or "organization").strip()) or "organization"
    return cleaned[:60]


@router.get("/me/volunteers/export-full")
def export_volunteers_full(
    format: str = "xlsx",
    current_user: dict = Depends(require_approved_org_admin),
):
    """Full-profile snapshot export of an organization's volunteers (xlsx | csv).

    Distinct from /me/volunteers/export (analytics-flat for Power BI). This endpoint
    returns the complete volunteer profile model for in-Excel review and analysis.
    """
    fmt = (format or "xlsx").lower()
    if fmt not in ("xlsx", "csv"):
        raise HTTPException(400, "format must be 'xlsx' or 'csv'")

    with get_db() as db:
        org_id = _get_my_org_id(db, current_user["id"])
        org = dict_row(db.execute("SELECT name FROM organizations WHERE id = ?", (org_id,)).fetchone())
        rows = _fetch_export_rows(db, org_id)

    shaped = [_shape_row(r) for r in rows]
    org_slug = _safe_filename(org["name"] if org else "organization")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    base_name = f"volunteers_{org_slug}_{today}"

    if fmt == "csv":
        output = io.StringIO()
        # UTF-8 BOM so Excel renders Arabic text correctly when opening CSVs.
        output.write("﻿")
        writer = csv.DictWriter(output, fieldnames=_EXPORT_COLUMNS, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in shaped:
            writer.writerow(row)
        return Response(
            content=output.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{base_name}.csv"',
                "X-Export-Count": str(len(shaped)),
            },
        )

    # xlsx path
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Volunteers"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="16A34A")
    header_align = Alignment(horizontal="left", vertical="center")

    ws.append(_EXPORT_COLUMNS)
    for col_idx, _ in enumerate(_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    wrap_align = Alignment(wrap_text=True, vertical="top")
    for row in shaped:
        excel_row = []
        for col in _EXPORT_COLUMNS:
            val = row.get(col, "")
            if col in _TEXT_COLUMNS and val != "":
                # Force text storage to preserve leading zeros in phone / national_id.
                val = str(val)
            excel_row.append(val)
        ws.append(excel_row)

        r = ws.max_row
        for col_idx, col in enumerate(_EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=r, column=col_idx)
            if col in _TEXT_COLUMNS:
                cell.number_format = "@"
            if col in _WIDE_WRAP_COLUMNS:
                cell.alignment = wrap_align

    # Column widths: roomy for wide/wrap columns, autosized-ish for the rest.
    for col_idx, col in enumerate(_EXPORT_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        if col in _WIDE_WRAP_COLUMNS:
            ws.column_dimensions[letter].width = 40
        elif col in {"email", "name", "supervisor_name", "university_name", "faculty"}:
            ws.column_dimensions[letter].width = 28
        elif col in {"joined_at", "date_of_birth"}:
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = 16

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{base_name}.xlsx"',
            "X-Export-Count": str(len(shaped)),
        },
    )


# ── Public listing ────────────────────────────────────────────────────────────

@router.get("")
def list_organizations():
    with get_db() as db:
        orgs = dict_rows(db.execute(
            "SELECT id, name, description, category, color, secondary_color, initials, founded, "
            "founded_year, location, org_type, logo_url, website "
            "FROM organizations WHERE status = 'approved' OR status IS NULL"
        ).fetchall())

        for org in orgs:
            oid = org["id"]
            org["total_volunteers"] = db.execute(
                "SELECT COUNT(*) as c FROM org_volunteers WHERE org_id = ? AND status = 'Active'", (oid,)
            ).fetchone()["c"]
            org["pending_requests"] = db.execute(
                "SELECT COUNT(*) as c FROM org_volunteers WHERE org_id = ? AND status = 'Pending'", (oid,)
            ).fetchone()["c"]
            org["active_activities"] = db.execute(
                "SELECT COUNT(*) as c FROM events WHERE org_id = ? AND status IN ('Active', 'Upcoming')", (oid,)
            ).fetchone()["c"]

        return {"organizations": orgs}


@router.get("/{org_id}")
def get_organization(org_id: int):
    with get_db() as db:
        org = dict_row(db.execute("SELECT * FROM organizations WHERE id = ?", (org_id,)).fetchone())
        if not org:
            raise HTTPException(404, "Organization not found")

        volunteers = dict_rows(db.execute(
            "SELECT v.id, v.name, v.email, v.phone, v.status, ov.department, ov.joined_date "
            "FROM volunteers v JOIN org_volunteers ov ON v.id = ov.volunteer_id WHERE ov.org_id = ?",
            (org["id"],),
        ).fetchall())

        supervisors = dict_rows(db.execute(
            "SELECT * FROM supervisors WHERE org_id = ?", (org["id"],)
        ).fetchall())

        events = dict_rows(db.execute(
            "SELECT * FROM events WHERE org_id = ? ORDER BY date DESC", (org["id"],)
        ).fetchall())

        return {**org, "volunteers": volunteers, "supervisors": supervisors, "events": events}


@router.get("/{org_id}/members")
def get_org_members(org_id: int, current_user: dict = Depends(get_current_user)):
    with get_db() as db:
        volunteers = dict_rows(db.execute(
            "SELECT v.*, ov.department, ov.status as org_status, ov.joined_date, s.name as supervisor_name "
            "FROM volunteers v JOIN org_volunteers ov ON v.id = ov.volunteer_id "
            "LEFT JOIN supervisors s ON ov.supervisor_id = s.id WHERE ov.org_id = ?",
            (org_id,),
        ).fetchall())

        supervisors = dict_rows(db.execute(
            "SELECT * FROM supervisors WHERE org_id = ?", (org_id,)
        ).fetchall())

        return {"volunteers": volunteers, "supervisors": supervisors}


@router.post("/{org_id}/join", status_code=201)
def join_organization(org_id: int, current_user: dict = Depends(require_roles("volunteer"))):
    with get_db() as db:
        vol = dict_row(db.execute(
            "SELECT id, governorate, city FROM volunteers WHERE user_id = ?", (current_user["id"],)
        ).fetchone())
        if not vol:
            raise HTTPException(404, "Volunteer profile not found")

        existing = dict_row(db.execute(
            "SELECT id FROM org_volunteers WHERE org_id = ? AND volunteer_id = ?",
            (org_id, vol["id"]),
        ).fetchone())
        if existing:
            raise HTTPException(409, "Already a member of this organization")

        _now = _utcnow()
        db.execute(
            "INSERT INTO org_volunteers "
            "(org_id, volunteer_id, status, is_active, join_source, "
            " governorate_snapshot, city_snapshot, joined_at) "
            "VALUES (?, ?, 'Pending', ?, 'website', ?, ?, ?)",
            (org_id, vol["id"], _is_active("Pending"),
             vol.get("governorate") or "", vol.get("city") or "", _now),
        )
        return {"message": "Application submitted"}


@router.put("/{org_id}/members/{vol_id}/approve")
def approve_org_member(
    org_id: int,
    vol_id: int,
    body: dict,
    current_user: dict = Depends(require_roles("org_admin")),
):
    with get_db() as db:
        db.execute(
            "UPDATE org_volunteers SET status = 'Active', is_active = ?, supervisor_id = ?, department = ? "
            "WHERE org_id = ? AND volunteer_id = ?",
            (_is_active("Active"), body.get("supervisor_id"), body.get("department"), org_id, vol_id),
        )
        return {"message": "Volunteer approved"}


@router.put("/{org_id}/members/{vol_id}/reject")
def reject_org_member(
    org_id: int,
    vol_id: int,
    current_user: dict = Depends(require_roles("org_admin")),
):
    with get_db() as db:
        db.execute(
            "DELETE FROM org_volunteers WHERE org_id = ? AND volunteer_id = ? AND status = 'Pending'",
            (org_id, vol_id),
        )
        return {"message": "Request rejected"}


def _assert_org_access(db, org_id: int, user_id: int):
    org = dict_row(db.execute(
        "SELECT id, name FROM organizations WHERE id = ? AND admin_user_id = ?",
        (org_id, user_id),
    ).fetchone())
    if not org:
        raise HTTPException(403, "Not authorized for this organization")
    return org


@router.post("/{org_id}/import-csv/preview")
def preview_import_csv(
    org_id: int,
    body: ImportCSVBody,
    current_user: dict = Depends(require_roles("org_admin")),
):
    """Dry-run: parse and validate CSV, return per-row status + summary without writing."""
    if not body.csv.strip():
        raise HTTPException(400, "CSV content is required")
    if body.duplicate_strategy not in DUPLICATE_STRATEGIES:
        raise HTTPException(400, f"duplicate_strategy must be one of: {', '.join(DUPLICATE_STRATEGIES)}")

    with get_db() as db:
        _assert_org_access(db, org_id, current_user["id"])

        reader      = csv.DictReader(io.StringIO(body.csv.strip()))
        result_rows = []
        seen_emails = set()

        for i, raw_row in enumerate(reader, start=2):
            cleaned, err = _validate_csv_row(raw_row, i)

            # Build a display row regardless of validity
            name  = (raw_row.get("name")  or raw_row.get("Name")  or "").strip()
            email = (raw_row.get("email") or raw_row.get("Email") or "").strip().lower()
            phone = (raw_row.get("phone") or "").strip()
            city  = (raw_row.get("city")  or raw_row.get("governorate") or "").strip()
            skills_raw = (raw_row.get("skills") or "").strip()

            if err:
                result_rows.append({
                    "name": name, "email": email, "phone": phone, "city": city,
                    "skills": skills_raw, "status": "error", "reason": err["reason"],
                })
                continue

            if cleaned["email"] in seen_emails:
                result_rows.append({
                    "name": name, "email": email, "phone": phone, "city": city,
                    "skills": skills_raw, "status": "error",
                    "reason": f"Duplicate email in CSV: {email}",
                })
                continue

            seen_emails.add(cleaned["email"])

            # DB lookup to determine status
            user_row = dict_row(db.execute(
                "SELECT id FROM users WHERE email = ?", (cleaned["email"],)
            ).fetchone())

            if not user_row:
                row_status = "new"
            else:
                vol_row = dict_row(db.execute(
                    "SELECT id FROM volunteers WHERE user_id = ?", (user_row["id"],)
                ).fetchone())
                vol_id = vol_row["id"] if vol_row else None
                if vol_id:
                    in_org = db.execute(
                        "SELECT id FROM org_volunteers WHERE org_id = ? AND volunteer_id = ?",
                        (org_id, vol_id),
                    ).fetchone()
                    row_status = "existing"  # covers both already-in and will-link
                else:
                    row_status = "existing"

            result_rows.append({
                "name": cleaned["name"], "email": cleaned["email"],
                "phone": cleaned["phone"], "city": cleaned["city"],
                "skills": ", ".join(cleaned["skills"]),
                "status": row_status,
            })

    summary = {
        "new":      sum(1 for r in result_rows if r["status"] == "new"),
        "existing": sum(1 for r in result_rows if r["status"] == "existing"),
        "errors":   sum(1 for r in result_rows if r["status"] == "error"),
    }
    return {"summary": summary, "rows": result_rows}


@router.post("/{org_id}/import-csv")
def import_volunteers_csv(
    org_id: int,
    body: ImportCSVBody,
    current_user: dict = Depends(require_roles("org_admin")),
):
    """Bulk-import volunteers from CSV.

    Extended headers: name, email, phone, city, skills, department, role, source
    New users are created without a password and receive a secure invite link.
    Existing users are linked to the organization directly.
    """
    if not body.csv.strip():
        raise HTTPException(400, "CSV content is required")
    if body.duplicate_strategy not in DUPLICATE_STRATEGIES:
        raise HTTPException(400, f"duplicate_strategy must be one of: {', '.join(DUPLICATE_STRATEGIES)}")

    valid_rows, parse_errors, _csv_dupes, quality_warnings = _parse_csv(body.csv)

    invited_count  = 0
    linked_count   = 0
    skipped_count  = 0
    runtime_errors = []
    invite_links   = []

    with get_db() as db:
        org = _assert_org_access(db, org_id, current_user["id"])
        org_name = org["name"]

        for r in valid_rows:
            email = r["email"]
            try:
                existing_user = dict_row(db.execute(
                    "SELECT id, role, invite_token FROM users WHERE email = ?", (email,)
                ).fetchone())

                if existing_user:
                    user_id = existing_user["id"]
                    vol_row = dict_row(db.execute(
                        "SELECT id FROM volunteers WHERE user_id = ?", (user_id,)
                    ).fetchone())
                    vol_id = vol_row["id"] if vol_row else None

                    if vol_id is None:
                        cur = db.execute(
                            "INSERT INTO volunteers (user_id, name, email, phone, city, skills, status) "
                            "VALUES (?, ?, ?, ?, ?, ?, 'Active')",
                            (user_id, r["name"], email, r["phone"] or None, r["city"] or None,
                             json.dumps(r["skills"])),
                        )
                        vol_id = cur.lastrowid

                    in_org = db.execute(
                        "SELECT id FROM org_volunteers WHERE org_id = ? AND volunteer_id = ?",
                        (org_id, vol_id),
                    ).fetchone()

                    if in_org:
                        if body.duplicate_strategy == "skip_duplicates":
                            skipped_count += 1
                            continue
                        elif body.duplicate_strategy == "update_existing":
                            db.execute(
                                "UPDATE org_volunteers "
                                "SET department = ?, source = ?, join_source = ?, channel_detail = ? "
                                "WHERE org_id = ? AND volunteer_id = ?",
                                (r["department"] or None, r["source"],
                                 r["join_source"], r["channel_detail"],
                                 org_id, vol_id),
                            )
                            skipped_count += 1
                            continue
                        # invite_anyway: fall through to re-invite logic below only if pending invite
                        if existing_user.get("invite_token"):
                            token = secrets.token_urlsafe(32)
                            expires = (datetime.now(timezone.utc) + timedelta(hours=48)).isoformat()
                            db.execute(
                                "UPDATE users SET invite_token = ?, invite_expires_at = ? WHERE id = ?",
                                (token, expires, user_id),
                            )
                            invite_links.append({"email": email, "link": _make_invite_link(token)})
                        skipped_count += 1
                        continue

                    # Existing user, not yet in this org — link them directly
                    _now = _utcnow()
                    db.execute(
                        "INSERT INTO org_volunteers "
                        "(org_id, volunteer_id, status, is_active, department, source, join_source, "
                        " channel_detail, governorate_snapshot, city_snapshot, joined_at) "
                        "VALUES (?, ?, 'Active', ?, ?, 'existing_user', ?, ?, ?, ?, ?)",
                        (org_id, vol_id, _is_active("Active"), r["department"] or None,
                         r["join_source"], r["channel_detail"],
                         r["governorate_snap"], r["city_snap"], _now),
                    )
                    linked_count += 1

                else:
                    # Brand-new user — create without password, issue invite token
                    token   = secrets.token_urlsafe(32)
                    expires = (datetime.now(timezone.utc) + timedelta(hours=48)).isoformat()
                    cur = db.execute(
                        "INSERT INTO users (email, password, role, invite_token, invite_expires_at) "
                        "VALUES (?, '', 'volunteer', ?, ?)",
                        (email, token, expires),
                    )
                    user_id = cur.lastrowid
                    cur = db.execute(
                        "INSERT INTO volunteers (user_id, name, email, phone, city, skills, status) "
                        "VALUES (?, ?, ?, ?, ?, ?, 'Pending')",
                        (user_id, r["name"], email, r["phone"] or None, r["city"] or None,
                         json.dumps(r["skills"])),
                    )
                    vol_id = cur.lastrowid
                    _now = _utcnow()
                    db.execute(
                        "INSERT INTO org_volunteers "
                        "(org_id, volunteer_id, status, is_active, department, source, join_source, "
                        " channel_detail, governorate_snapshot, city_snapshot, joined_at) "
                        "VALUES (?, ?, 'Pending', ?, ?, 'invite', ?, ?, ?, ?, ?)",
                        (org_id, vol_id, _is_active("Pending"), r["department"] or None,
                         r["join_source"], r["channel_detail"],
                         r["governorate_snap"], r["city_snap"], _now),
                    )
                    invite_links.append({"email": email, "link": _make_invite_link(token)})
                    invited_count += 1

            except Exception as e:
                runtime_errors.append({"row": email, "reason": str(e)})

    all_errors = parse_errors + runtime_errors
    success_count = invited_count + linked_count

    return {
        "totalRows":       len(valid_rows) + len(parse_errors),
        "successCount":    success_count,
        "invitedCount":    invited_count,
        "linkedCount":     linked_count,
        "skippedCount":    skipped_count,
        "errorCount":      len(all_errors),
        "errors":          all_errors,
        "inviteLinks":     invite_links,
        # Data-quality report: rows where join_source was invalid and coerced to "other".
        "qualityWarnings": quality_warnings,
    }


@router.delete("/{org_id}/members/{vol_id}")
def remove_org_member(
    org_id: int,
    vol_id: int,
    current_user: dict = Depends(require_roles("org_admin")),
):
    with get_db() as db:
        db.execute(
            "DELETE FROM org_volunteers WHERE org_id = ? AND volunteer_id = ?",
            (org_id, vol_id),
        )
        return {"message": "Volunteer removed from organization"}
